import pandas as pd
from openpyxl import load_workbook

# Пути к файлам
input_file = 'GTO_GAS_122023.xls'
output_file = 'вывод.xlsx'

# Получаем список всех листов, соответствующих шаблону «УКПГ-*»
xls = pd.ExcelFile(input_file)
sheets = [sheet for sheet in xls.sheet_names if sheet.startswith('УКПГ-')]

# Загружаем существующий файл Excel
book = load_workbook(output_file)

# Получаем имя первого листа
first_sheet_name = book.sheetnames[0]

# Определяем, где начинать запись (максимальная строка)
start_row = book[first_sheet_name].max_row

result_data = pd.DataFrame()

for sheet in sheets:
    # Читаем данные из листа
    df = pd.read_excel(xls, sheet_name=sheet, header=None)

    for i in range(10, len(df)):  # Индекс 9 соответствует 10-й строке
        serial_number = df.iloc[i, 0]  # № ПП
        # print(serial_number)
        well_number = df.iloc[i, 1]  # Номер скважины
        indic = df.iloc[i, 2]  # Инжикатор
        object = df.iloc[i, 3]  # Объект
        date_begin = df.iloc[i, 4]  #
        r_gol = df.iloc[i, 5]
        r_ztr = df.iloc[i, 6]

        debet_gas = df.iloc[i, 7]  # Дебет газа
        gas_production_per_month = df.iloc[i, 8]  # добыча газа за месяц
        gas_production_begin_year = df.iloc[i, 9]  # добыча газа с начала года
        gas_production_begin_exploitation = df.iloc[i, 10]  # добыча газа с начала экспл

        debet_condensate = df.iloc[i, 11]  # дебит конденсата
        production_condensate_month = df.iloc[i, 12]  # добыча конденсата за месяц
        production_condensate_begin_year = df.iloc[i, 13]  # добыча конденсата с начала года
        production_condensate_begin_exploitation = df.iloc[i, 14]  # добыча конденсата с начала экспл

        debet_water = df.iloc[i, 15]  # дебит конденсата
        production_water_month = df.iloc[i, 16]  # добыча воды за месяц
        production_water_begin_year = df.iloc[i, 17]  # добыча воды с начала года
        production_water_begin_exploitation = df.iloc[i, 18]  # добыча воды с начала экспл

        calendar_time = df.iloc[i, 19]  # календарн время
        days_exploitation = df.iloc[i, 20]  # дни экспл
        fact_time = df.iloc[i, 21]  # факт время
        downtime = df.iloc[i, 22]  # простои

        value_in_second_row = df.iloc[1, 13]  # Получаем значение во второй строке и колонке N
        date_str = value_in_second_row.split()
        month, year = date_str[0], date_str[1]  # Месяц и год

        if isinstance(serial_number, (int, float)) and serial_number == i - 9:  # Проверяем, является ли это порядковым номером
            # Создаем временный DataFrame для добавления
            temp_df = pd.DataFrame({
                0: [sheet],
                1: [month],
                2: [year],
                3: [serial_number],
                4: [well_number],
                5: [indic],
                6: [object],
                7: [date_begin],
                8: [r_gol],
                9: [r_ztr],
                10: [debet_gas],
                11: [gas_production_per_month],
                12: [gas_production_begin_year],
                13: [gas_production_begin_exploitation],
                14: [debet_condensate],
                15: [production_condensate_month],
                16: [production_condensate_begin_year],
                17: [production_condensate_begin_exploitation],
                18: [debet_water],
                19: [production_water_month],
                20: [production_water_begin_year],
                21: [production_water_begin_exploitation],
                22: [calendar_time],
                23: [days_exploitation],
                24: [fact_time],
                25: [downtime]
            })
            result_data = pd.concat([result_data, temp_df], ignore_index=True)  # Объединяем

# Записываем названия листов в существующий файл, добавляя данные без заголовков
with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
    result_data.to_excel(writer, index=False, header=False, startrow=start_row, startcol=0, sheet_name=first_sheet_name)



